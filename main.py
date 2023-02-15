import paramiko
import openpyxl
import time


# 读取Excel表格中的设备信息
def read_device_info_from_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    device_info_list = []
    for i in range(2, sheet.max_row + 1):
        device_info = {}
        device_info['IP'] = sheet.cell(row=i, column=1).value
        device_info['username'] = sheet.cell(row=i, column=2).value
        device_info['password'] = sheet.cell(row=i, column=3).value
        device_info['equipment name'] = sheet.cell(row=i, column=4).value
        device_info_list.append(device_info)
    return device_info_list


# 连接设备并执行命令
def ssh_to_device_and_exec_cmd(device_info, cmd):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(device_info['IP'], 22, device_info['username'], device_info['password'])
    stdin, stdout, stderr = ssh.exec_command(cmd)
    result = stdout.read().decode(encoding='gbk')
    ssh.close()
    return result


# 将结果保存到本地
# def save_result_to_local(result, file_name):
# with open(file_name, 'w') as f:
# f.write(result)

# 将结果保存到指定文件夹下面
def save_result_to_local(result, file_name):
    with open('C:/Users/admin/Desktop/devices/backup/' + file_name, 'w') as f:
        f.write(result)


if __name__ == '__main__':
    while True:
        device_info_list = read_device_info_from_excel('devices.xlsx')
        for device_info in device_info_list:
            result = ssh_to_device_and_exec_cmd(device_info, 'dis cu')
            filename = device_info['equipment name'] + '-' + device_info['IP'] + '.txt'
            save_result_to_local(result, filename)
        time.sleep(3600)  # 每小时执行一次
